"""
core/utils.py
─────────────
Funções utilitárias de dados compartilhadas entre os três módulos.
Consolida as três implementações paralelas que existiam nos sistemas originais:
  - limpar_moeda / parse_br_float / para_float  →  moeda_para_float()
  - limpar_matricula                             →  limpar_matricula()
  - constantes de cor Excel                     →  COR_*, FONTE_*
"""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ════════════════════════════════════════════════════════════
# CONSTANTES DE FORMATAÇÃO EXCEL  (usadas nos três módulos)
# ════════════════════════════════════════════════════════════

COR_VERDE     = PatternFill("solid", fgColor="C6EFCE")
COR_VERMELHO  = PatternFill("solid", fgColor="FFC7CE")
COR_AMARELO   = PatternFill("solid", fgColor="FFEB9C")
COR_CABECALHO = PatternFill("solid", fgColor="1F4E79")
COR_RESUMO    = PatternFill("solid", fgColor="BDD7EE")

FONTE_CAB    = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
FONTE_BOLD   = Font(bold=True, name="Calibri", size=10)
FONTE_NORMAL = Font(name="Calibri", size=10)

BORDA = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center")
ALINHAMENTO_ESQ    = Alignment(horizontal="left",   vertical="center")


# ════════════════════════════════════════════════════════════
# FUNÇÕES DE LIMPEZA E CONVERSÃO
# ════════════════════════════════════════════════════════════

def moeda_para_float(valor) -> float:
    """
    Converte qualquer representação de valor monetário brasileiro para float.
    Trata: None, NaN, int, float, strings com R$, ponto-milhar e vírgula-decimal.

    Exemplos:
        "R$ 1.234,56" → 1234.56
        "1234,56"     → 1234.56
        "1234.56"     → 1234.56
        None / NaN    → 0.0
    """
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return 0.0 if pd.isna(valor) else float(valor)

    v = str(valor).strip().replace("R$", "").strip()
    if not v:
        return 0.0

    # Formato BR: ponto como milhar, vírgula como decimal
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        v = v.replace(",", ".")

    try:
        return float(v)
    except ValueError:
        return 0.0


def limpar_str(valor) -> str:
    """Remove espaços extras, tabulações e quebras de linha de qualquer valor."""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return str(valor).strip().replace("\n", " ").replace("\t", " ")


def limpar_matricula(valor) -> str:
    """
    Normaliza matrícula para string sem decimais espúrios.
    Exemplo: 276.0 → "276"
    """
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    v = str(valor).strip()
    if v.endswith(".0"):
        v = v[:-2]
    return v


def formatar_moeda_br(valor: float) -> str:
    """
    Formata float para exibição em padrão BR.
    Exemplo: 1234.5 → "R$ 1.234,50"
    """
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def ler_csv_com_fallback(arquivo, **kwargs) -> pd.DataFrame:
    """
    Lê CSV tentando UTF-8 primeiro. Vários sistemas de RH/folha exportam
    em Windows-1252/Latin-1 (acentos em nomes e cargos), o que quebra o
    parser padrão do pandas com UnicodeDecodeError — cai para latin1
    nesse caso, que aceita qualquer byte sem erro.
    """
    try:
        return pd.read_csv(arquivo, encoding="utf-8-sig", **kwargs)
    except UnicodeDecodeError:
        arquivo.seek(0)
        return pd.read_csv(arquivo, encoding="latin1", **kwargs)
