"""
services/adiantamento.py
────────────────────────
Lógica de leitura, cálculo de proporção de férias e conferência de adiantamentos.
"""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Importando as funções globais e constantes do core
from core.utils import (
    moeda_para_float, ler_csv_com_fallback,
    COR_VERDE, COR_VERMELHO, COR_CABECALHO, COR_RESUMO,
    FONTE_CAB, FONTE_BOLD, FONTE_NORMAL, BORDA
)

# Cores específicas herdadas da lógica original do adiantamento
COR_AZUL_NOVO = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
COR_CINZA_OPC = PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid')

COLUNAS_SAIDA = [
    "Matrícula", "Funcionário", "Data de Admissão", "Categoria", "Salário Base",
    "Adiantamento Mês Anterior", "Adiantamento Mês Atual", "Diferença Entre Meses",
    "Status", "Motivo / Erros"
]

def nome_arquivo(arquivo) -> str:
    return str(getattr(arquivo, "name", getattr(arquivo, "filename", ""))).lower()

def reposicionar(arquivo):
    try:
        arquivo.seek(0)
    except Exception:
        pass

def ler_planilha(arquivo, header=None) -> pd.DataFrame:
    reposicionar(arquivo)
    if nome_arquivo(arquivo).endswith(".csv"):
        return ler_csv_com_fallback(arquivo, header=header, sep=None, engine="python")
    return pd.read_excel(arquivo, header=header)

def verificar_optante(val) -> bool:
    v = str(val).strip().lower()
    if v in ['0', '0.0', 'não', 'nao', 'n', 'false']:
        return False
    return True

def tem_direito_adiantamento(mes: int, ano: int, data_adm) -> bool:
    if pd.isna(data_adm): return True
    data_ref_inicio = pd.Timestamp(year=ano, month=mes, day=1)
    if data_adm < data_ref_inicio: return True
    if data_adm.year == ano and data_adm.month == mes:
        return data_adm.day <= 6
    return False

def calcular_dias_ferias_no_mes(mes: int, ano: int, data_ini, data_fim) -> int:
    if pd.isna(data_ini) or pd.isna(data_fim): return 0
    primeiro_dia = pd.Timestamp(year=ano, month=mes, day=1)
    ultimo_dia_real = pd.Timestamp(year=ano, month=mes, day=1) + pd.offsets.MonthEnd(0)
    
    if data_fim < primeiro_dia or data_ini > ultimo_dia_real:
        return 0
        
    overlap_start = max(primeiro_dia, data_ini)
    overlap_end = min(ultimo_dia_real, data_fim)
    
    d_start = overlap_start.day
    if d_start == 31: d_start = 30
    
    d_end = overlap_end.day
    if d_end == 31: d_end = 30
    
    if overlap_start == primeiro_dia and overlap_end == ultimo_dia_real:
        return 30
        
    dias = d_end - d_start + 1
    return max(0, dias)

def executar_conferencia_adiantamento(file_eventos, file_ativos, file_ferias) -> tuple[list[dict], dict]:
    df_ev_raw = ler_planilha(file_eventos, header=None)
    df_at_raw = ler_planilha(file_ativos, header=None)
    df_fe_raw = ler_planilha(file_ferias, header=None)

    df_ev = df_ev_raw.iloc[2:, [1, 2, 11, 12, 13, 14, 16]].copy()
    df_ev.columns = ['Matricula', 'Nome_Ev', 'Mes', 'Ano', 'Cod_Evento', 'Nome_Evento', 'Valor_Provento']
    
    df_at = df_at_raw.iloc[2:, [1, 4, 44, 51, 62, 64]].copy()
    df_at.columns = ['Matricula', 'Nome', 'Data_Admissao', 'Categoria', 'Salario', 'Opta_Adiantamento']
    
    df_fe = df_fe_raw.iloc[2:, [4, 6, 14, 15]].copy()
    df_fe.columns = ['Matricula', 'Nome_Fe', 'Data_Ini_Ferias', 'Data_Fim_Ferias']

    def limpa_mat(val):
        if pd.isna(val): return ""
        v = str(val).strip()
        if v.endswith('.0'): v = v[:-2]
        return v

    df_ev['Matricula'] = df_ev['Matricula'].apply(limpa_mat)
    df_at['Matricula'] = df_at['Matricula'].apply(limpa_mat)
    df_fe['Matricula'] = df_fe['Matricula'].apply(limpa_mat)
    
    df_ev = df_ev[df_ev['Matricula'] != ""]
    df_at = df_at[df_at['Matricula'] != ""]
    df_fe = df_fe[df_fe['Matricula'] != ""]

    df_ev['Valor_Provento'] = df_ev['Valor_Provento'].apply(moeda_para_float)
    df_at['Salario'] = df_at['Salario'].apply(moeda_para_float)
    df_ev['Mes'] = pd.to_numeric(df_ev['Mes'], errors='coerce').fillna(0).astype(int)
    df_ev['Ano'] = pd.to_numeric(df_ev['Ano'], errors='coerce').fillna(0).astype(int)
    df_ev['Cod_Evento'] = pd.to_numeric(df_ev['Cod_Evento'], errors='coerce').fillna(0).astype(int)
    
    df_at['Data_Admissao'] = pd.to_datetime(df_at['Data_Admissao'], errors='coerce', dayfirst=True)
    df_fe['Data_Ini_Ferias'] = pd.to_datetime(df_fe['Data_Ini_Ferias'], errors='coerce', dayfirst=True)
    df_fe['Data_Fim_Ferias'] = pd.to_datetime(df_fe['Data_Fim_Ferias'], errors='coerce', dayfirst=True)

    df_ev = df_ev[df_ev['Mes'] > 0]
    meses_disponiveis = sorted(df_ev['Mes'].unique())
    if len(meses_disponiveis) < 2:
        raise ValueError("A planilha de EVENTOS precisa de dados de 2 meses distintos (Anterior e Atual).")
        
    mes_anterior = int(meses_disponiveis[0])
    mes_atual = int(meses_disponiveis[1])

    df_ev_100 = df_ev[df_ev['Cod_Evento'] == 100]
    tot_ant_global = df_ev_100[df_ev_100['Mes'] == mes_anterior]['Valor_Provento'].sum()
    tot_atu_global = df_ev_100[df_ev_100['Mes'] == mes_atual]['Valor_Provento'].sum()

    try: ano_ant = int(df_ev[df_ev['Mes'] == mes_anterior]['Ano'].mode()[0])
    except: ano_ant = 2026
    try: ano_atu = int(df_ev[df_ev['Mes'] == mes_atual]['Ano'].mode()[0])
    except: ano_atu = 2026

    matriculas_evento_invalido = df_ev[df_ev['Cod_Evento'] != 100]['Matricula'].unique()

    pivot_ev = df_ev_100.pivot_table(index='Matricula', columns='Mes', values='Valor_Provento', aggfunc='sum').reset_index()
    if mes_anterior not in pivot_ev.columns: pivot_ev[mes_anterior] = 0.0
    if mes_atual not in pivot_ev.columns: pivot_ev[mes_atual] = 0.0
    pivot_ev.rename(columns={mes_anterior: 'Valor_Mes_Anterior', mes_atual: 'Valor_Mes_Atual'}, inplace=True)

    report = pd.merge(df_at, pivot_ev, on='Matricula', how='left')
    report['Valor_Mes_Anterior'] = report['Valor_Mes_Anterior'].fillna(0.0)
    report['Valor_Mes_Atual'] = report['Valor_Mes_Atual'].fillna(0.0)

    resultados = []
    
    for _, row in report.iterrows():
        erros = []
        matricula = row['Matricula']
        categoria = str(row['Categoria']).strip()
        salario = row['Salario']
        val_ant = row['Valor_Mes_Anterior']
        val_atu = row['Valor_Mes_Atual']
        data_adm = row['Data_Admissao']
        
        is_aprendiz = "Aprendiz (Lei 10.097/2000)" in categoria
        optante = verificar_optante(row['Opta_Adiantamento'])
        direito_ant = tem_direito_adiantamento(mes_anterior, ano_ant, data_adm)
        direito_atu = tem_direito_adiantamento(mes_atual, ano_atu, data_adm)

        ferias_func = df_fe[(df_fe['Matricula'] == matricula)]
        dias_fe_ant = 0
        dias_fe_atu = 0
        
        for _, v_row in ferias_func.iterrows():
            dias_fe_ant += calcular_dias_ferias_no_mes(mes_anterior, ano_ant, v_row['Data_Ini_Ferias'], v_row['Data_Fim_Ferias'])
            dias_fe_atu += calcular_dias_ferias_no_mes(mes_atual, ano_atu, v_row['Data_Ini_Ferias'], v_row['Data_Fim_Ferias'])
            
        if dias_fe_ant > 30: dias_fe_ant = 30
        if dias_fe_atu > 30: dias_fe_atu = 30
        
        dias_trab_ant = 30 - dias_fe_ant
        dias_trab_atu = 30 - dias_fe_atu

        if dias_trab_ant < 15: esperado_ant = 0.0
        else: esperado_ant = round((salario * 0.40) / 30 * dias_trab_ant, 2)
            
        if dias_trab_atu < 15: esperado_atu = 0.0
        else: esperado_atu = round((salario * 0.40) / 30 * dias_trab_atu, 2)

        if not optante:
            status_final = "Sem adiantamento (opcional)"
            descricao_erro = "Funcionário não optante (Coluna BM = 0)"
        elif is_aprendiz:
            if val_atu > 0 or val_ant > 0: erros.append("Aprendiz não deve receber adiantamento")
            status_final = "Errado" if erros else "Certo"
            descricao_erro = " - ".join(erros) if erros else "Correto (Aprendiz zerado)"
        else:
            if matricula in matriculas_evento_invalido: 
                erros.append("Contém evento diferente de 100")
                
            if not direito_atu:
                if val_atu > 0: erros.append("Recebeu indevidamente (Admitido após o dia 6)")
            else:
                if dias_trab_atu >= 15:
                    if val_atu == 0: 
                        erros.append("Falta adiantamento no mês atual")
                    elif val_atu > 0 and abs(round(val_atu, 2) - esperado_atu) > 0.02:
                        if dias_trab_atu < 30:
                            erros.append(f"Cálculo de Férias incorreto (Esperado: R$ {esperado_atu:.2f} p/ {dias_trab_atu} dias trab.)")
                        else:
                            erros.append(f"Cálculo incorreto (Esperado: R$ {esperado_atu:.2f})")
                else:
                    if val_atu > 0:
                        erros.append(f"Recebeu indevidamente (Trabalhou apenas {dias_trab_atu} dias no mês)")

            # Não comparamos mais o valor do mês atual contra o mês anterior:
            # o cálculo de "esperado_atu" acima já garante que o valor pago bate
            # com os 40% do salário atual (considerando os dias de férias do mês).
            # Se bateu, está certo — mesmo que o mês anterior seja diferente por
            # causa de um reajuste salarial no meio do caminho.

            if len(erros) > 0:
                status_final = "Errado"
                descricao_erro = " - ".join(erros)
            else:
                if not direito_atu:
                    status_final = "Não tem direito (admitido após dia 6)"
                    data_str = data_adm.strftime('%d/%m/%Y') if pd.notna(data_adm) else "Recente"
                    descricao_erro = f"Isento - Admitido em {data_str}"
                elif not direito_ant and direito_atu:
                    status_final = "Funcionário Novo"
                    descricao_erro = "Primeiro adiantamento (Isento de comp. c/ mês anterior)"
                elif dias_trab_atu < 15:
                    status_final = "Certo (Férias)"
                    descricao_erro = f"Isento de adiantamento ({dias_trab_atu} dias trab. no mês)"
                elif dias_trab_atu < 30:
                    status_final = "Certo (Férias)"
                    descricao_erro = f"Proporcional correto ({dias_trab_atu} dias trab. / {dias_fe_atu} dias férias)"
                else:
                    status_final = "Certo"
                    descricao_erro = "Sem divergências"
        
        data_adm_formatada = data_adm.strftime('%d/%m/%Y') if pd.notna(data_adm) else ""

        resultados.append({
            "Matrícula": matricula,
            "Funcionário": row['Nome'],
            "Data de Admissão": data_adm_formatada,
            "Categoria": categoria,
            "Salário Base": salario,
            "Adiantamento Mês Anterior": val_ant,
            "Adiantamento Mês Atual": val_atu,
            "Diferença Entre Meses": val_atu - val_ant,
            "Status": status_final,
            "Motivo / Erros": descricao_erro
        })

    meta = {
        "mes_ant": mes_anterior,
        "mes_atu": mes_atual,
        "tot_ant": float(tot_ant_global),
        "tot_atu": float(tot_atu_global),
        "total_ativos": len(resultados),
        "total_corretos": sum(1 for r in resultados if "Certo" in r["Status"]),
        "total_errados": sum(1 for r in resultados if r["Status"] == "Errado"),
        "total_isentos": sum(1 for r in resultados if r["Status"] in ["Funcionário Novo", "Sem adiantamento (opcional)", "Não tem direito (admitido após dia 6)"])
    }

    return resultados, meta

def gerar_excel_adiantamento(resultados: list[dict], meta: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFERÊNCIA"

    for ci, titulo in enumerate(COLUNAS_SAIDA, 1):
        c = ws.cell(row=1, column=ci, value=titulo)
        c.fill, c.font, c.border = COR_CABECALHO, FONTE_CAB, BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ri, reg in enumerate(resultados, 2):
        for ci, col in enumerate(COLUNAS_SAIDA, 1):
            valor = reg.get(col, "")
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font, c.border, c.alignment = FONTE_NORMAL, BORDA, Alignment(vertical="center")
            
            if col in ["Salário Base", "Adiantamento Mês Anterior", "Adiantamento Mês Atual", "Diferença Entre Meses"]:
                c.number_format = "R$ #,##0.00"
                c.alignment = Alignment(horizontal="right")
            elif col in ["Matrícula", "Data de Admissão", "Status"]:
                c.alignment = Alignment(horizontal="center")

        status = reg.get("Status", "")
        if "Certo" in status: fill_cor = COR_VERDE
        elif status == "Errado": fill_cor = COR_VERMELHO
        elif status == "Funcionário Novo": fill_cor = COR_AZUL_NOVO
        else: fill_cor = COR_CINZA_OPC

        ws.cell(row=ri, column=9).fill = fill_cor

    ws.auto_filter.ref, ws.freeze_panes = ws.dimensions, "C2"
    
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    ws2 = wb.create_sheet("RESUMO")
    linhas_resumo = [
        (f"AUDITORIA DE ADIANTAMENTO — MÊS {meta['mes_atu']}", "", ""),
        ("", "", ""),
        ("INDICADOR", "QUANTIDADE", "VALOR GLOBAL"),
        ("Total de Funcionários Ativos", meta['total_ativos'], ""),
        ("✅ Empregados Corretos", meta['total_corretos'], ""),
        ("❌ Empregados com Erros/Divergências", meta['total_errados'], ""),
        ("⚠️ Isentos (Novos / Não Optantes)", meta['total_isentos'], ""),
        ("", "", ""),
        ("Faturamento Adiantamento Mês Anterior", "", meta['tot_ant']),
        ("Faturamento Adiantamento Mês Atual", "", meta['tot_atu']),
        ("Diferença Total Líquida (Mês Atual x Ant)", "", meta['tot_atu'] - meta['tot_ant']),
    ]

    for ri, tripla in enumerate(linhas_resumo, 1):
        a, b, c_val = tripla
        for ci, val in enumerate((a, b, c_val), 1):
            cel = ws2.cell(row=ri, column=ci, value=val)
            cel.border, cel.font = BORDA, FONTE_NORMAL
            cel.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == 3 and isinstance(val, (int, float)):
                cel.number_format = "R$ #,##0.00"

    ws2.merge_cells("A1:C1")
    ws2["A1"].fill, ws2["A1"].font, ws2["A1"].alignment = COR_CABECALHO, Font(color="FFFFFF", bold=True, name="Calibri", size=13), Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    for ci in range(1, 4): ws2.cell(row=3, column=ci).fill, ws2.cell(row=3, column=ci).font = COR_RESUMO, FONTE_BOLD

    ws2.column_dimensions["A"].width, ws2.column_dimensions["B"].width, ws2.column_dimensions["C"].width = 42, 16, 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def processar_dados(file_eventos, file_ativos, file_ferias):
    """Compatibilidade com a interface Streamlit original do módulo."""
    resultados, meta = executar_conferencia_adiantamento(file_eventos, file_ativos, file_ferias)
    df = pd.DataFrame(resultados).rename(columns={
        "Matrícula": "Matricula",
        "Funcionário": "Nome",
        "Salário Base": "Salario",
        "Adiantamento Mês Anterior": f"Adiantamento (Mês {meta['mes_ant']})",
        "Adiantamento Mês Atual": f"Adiantamento (Mês {meta['mes_atu']})",
    })
    return df, meta["mes_ant"], meta["mes_atu"], meta["tot_ant"], meta["tot_atu"]


def gerar_excel_formatado(df: pd.DataFrame, mes_ant: int, mes_atu: int) -> bytes:
    """Gera Excel a partir do DataFrame filtrado exibido no Streamlit."""
    df_export = df.rename(columns={
        "Matricula": "Matrícula",
        "Nome": "Funcionário",
        "Salario": "Salário Base",
        f"Adiantamento (Mês {mes_ant})": "Adiantamento Mês Anterior",
        f"Adiantamento (Mês {mes_atu})": "Adiantamento Mês Atual",
    })

    meta = {
        "mes_ant": mes_ant,
        "mes_atu": mes_atu,
        "tot_ant": float(df_export["Adiantamento Mês Anterior"].sum()) if "Adiantamento Mês Anterior" in df_export else 0.0,
        "tot_atu": float(df_export["Adiantamento Mês Atual"].sum()) if "Adiantamento Mês Atual" in df_export else 0.0,
        "total_ativos": len(df_export),
        "total_corretos": int(df_export["Status"].str.contains("Certo", na=False).sum()) if "Status" in df_export else 0,
        "total_errados": int((df_export["Status"] == "Errado").sum()) if "Status" in df_export else 0,
        "total_isentos": int(df_export["Status"].isin([
            "Funcionário Novo",
            "Sem adiantamento (opcional)",
            "Não tem direito (admitido após dia 6)",
        ]).sum()) if "Status" in df_export else 0,
    }

    registros = df_export.to_dict("records")
    return gerar_excel_adiantamento(registros, meta)
