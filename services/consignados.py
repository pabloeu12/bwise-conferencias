"""
services/consignados.py
───────────────────────
Lógica de conferência de empréstimos consignados.
"""

import io
import re
import unicodedata

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.utils import (
    BORDA,
    COR_CABECALHO,
    COR_RESUMO,
    COR_VERDE,
    COR_VERMELHO,
    FONTE_BOLD,
    FONTE_CAB,
    FONTE_NORMAL,
    ler_csv_com_fallback,
    limpar_matricula,
    moeda_para_float,
)

CODIGOS_FERIAS = [189, 190, 223]
CODIGOS_EMPRESTIMO = [652, 664, 665, 666, 667, 668, 669, 677, 678, 685, 694, 695]
LIMITE_DESCONTO_PERCENTUAL = 0.35

COLUNAS_SAIDA = [
    "Matrícula",
    "Nome do funcionário",
    "Base de Cálculo do INSS",
    "Valor do INSS",
    "Valor do IRRF",
    "Férias",
    "Base",
    "Limite de Desconto (35%)",
    "EMPRESTIMO (EMPREGA BRASIL)",
    "EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)",
    "Diferença (Emprega Brasil x Lista de Eventos)",
    "Diferença (Limite de Desconto x Lista de Eventos)",
    "Status",
    "Observação",
]


def normalizar_nome(nome) -> str:
    """Normaliza nomes para cruzamento entre relatórios distintos."""
    if pd.isna(nome):
        return ""
    texto = str(nome).upper()
    texto = "".join(
        ch for ch in unicodedata.normalize("NFD", texto)
        if unicodedata.category(ch) != "Mn"
    )
    texto = re.sub(r"[^A-Z\s]", " ", texto)
    return " ".join(texto.split())


def normalizar_cabecalho(valor) -> str:
    texto = "" if valor is None else str(valor).strip().lower()
    texto = "".join(
        ch for ch in unicodedata.normalize("NFD", texto)
        if unicodedata.category(ch) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def _nome_arquivo(arquivo) -> str:
    return str(getattr(arquivo, "name", getattr(arquivo, "filename", arquivo))).lower()


def _reposicionar(arquivo):
    try:
        arquivo.seek(0)
    except Exception:
        pass


def carregar_arquivo(arquivo, skiprows=0, header="infer") -> pd.DataFrame:
    """Carrega CSV/XLS/XLSX independente da origem do upload (FastAPI)."""
    _reposicionar(arquivo)
    nome = _nome_arquivo(arquivo)
    if nome.endswith(".csv"):
        return ler_csv_com_fallback(arquivo, skiprows=skiprows, header=header, sep=None, engine="python")
    return pd.read_excel(arquivo, skiprows=skiprows, header=header)


def _coluna_por_alias(df: pd.DataFrame, aliases: list[str]) -> str:
    normalizadas = {normalizar_cabecalho(col): col for col in df.columns}
    aliases_norm = [normalizar_cabecalho(alias) for alias in aliases]

    for alias in aliases_norm:
        if alias in normalizadas:
            return normalizadas[alias]

    for alias in aliases_norm:
        for normalizada, original in normalizadas.items():
            if alias and alias in normalizada:
                return original

    raise ValueError(f"Não foi possível identificar coluna: {', '.join(aliases)}")


def executar_conferencia_consignados(arquivo_emprega, arquivo_recibos, arquivo_eventos) -> tuple[list[dict], dict]:
    df_emprega = carregar_arquivo(arquivo_emprega, header=0)
    col_nome_trab = _coluna_por_alias(df_emprega, ["nomeTrabalhador", "nome trabalhador", "nome"])
    col_valor = _coluna_por_alias(df_emprega, ["valorParcela", "valor parcela", "valor"])

    df_emprega["Nome Normalizado"] = df_emprega[col_nome_trab].apply(normalizar_nome)
    df_emprega[col_valor] = df_emprega[col_valor].apply(moeda_para_float)
    emprega_agrupado = df_emprega.groupby("Nome Normalizado", as_index=False)[col_valor].sum()
    emprega_agrupado.rename(columns={col_valor: "EMPRESTIMO (EMPREGA BRASIL)"}, inplace=True)

    df_recibo = carregar_arquivo(arquivo_recibos, skiprows=2, header=None)
    indices_recibo = {
        "Matrícula": 1,
        "Nome do funcionário": 2,
        "Base de Cálculo do INSS": 27,
        "Valor do INSS": 29,
        "Valor do IRRF": 42,
    }
    if df_recibo.shape[1] <= max(indices_recibo.values()):
        raise ValueError("A Lista de Recibo de Pagamento não possui as colunas esperadas.")

    df_base = pd.DataFrame({
        "Matrícula": df_recibo[indices_recibo["Matrícula"]].apply(limpar_matricula),
        "Nome do funcionário": df_recibo[indices_recibo["Nome do funcionário"]],
        "Base de Cálculo do INSS": df_recibo[indices_recibo["Base de Cálculo do INSS"]].apply(moeda_para_float),
        "Valor do INSS": df_recibo[indices_recibo["Valor do INSS"]].apply(moeda_para_float),
        "Valor do IRRF": df_recibo[indices_recibo["Valor do IRRF"]].apply(moeda_para_float),
    })
    df_base = df_base.dropna(subset=["Matrícula", "Nome do funcionário"])
    df_base = df_base[df_base["Matrícula"] != ""]
    df_base["Nome Normalizado"] = df_base["Nome do funcionário"].apply(normalizar_nome)

    df_eventos = carregar_arquivo(arquivo_eventos, header=1)
    try:
        col_nome_evento = _coluna_por_alias(df_eventos, ["Nome", "Nome do funcionário"])
    except ValueError:
        df_eventos = carregar_arquivo(arquivo_eventos, header=0)
        col_nome_evento = _coluna_por_alias(df_eventos, ["Nome", "Nome do funcionário"])

    col_cod_evento = _coluna_por_alias(df_eventos, ["Cód. Evento", "Cod Evento", "Código Evento"])
    col_valor_provento = _coluna_por_alias(df_eventos, ["Valor Provento", "Provento"])
    col_valor_desconto = _coluna_por_alias(df_eventos, ["Valor Desconto", "Desconto"])

    df_eventos["Nome Normalizado"] = df_eventos[col_nome_evento].apply(normalizar_nome)
    df_eventos[col_valor_provento] = df_eventos[col_valor_provento].apply(moeda_para_float)
    df_eventos[col_valor_desconto] = df_eventos[col_valor_desconto].apply(moeda_para_float)
    df_eventos["Cód. Evento Num"] = pd.to_numeric(df_eventos[col_cod_evento], errors="coerce")

    df_ferias = df_eventos[df_eventos["Cód. Evento Num"].isin(CODIGOS_FERIAS)]
    df_ferias_agrupado = df_ferias.groupby("Nome Normalizado", as_index=False)[col_valor_provento].sum()
    df_ferias_agrupado.rename(columns={col_valor_provento: "Férias"}, inplace=True)

    df_emp_folha = df_eventos[df_eventos["Cód. Evento Num"].isin(CODIGOS_EMPRESTIMO)]
    df_emp_folha_agrupado = df_emp_folha.groupby("Nome Normalizado", as_index=False)[col_valor_desconto].sum()
    df_emp_folha_agrupado.rename(
        columns={col_valor_desconto: "EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)"},
        inplace=True,
    )

    df_final = pd.merge(df_base, df_ferias_agrupado, on="Nome Normalizado", how="left")
    df_final = pd.merge(df_final, emprega_agrupado, on="Nome Normalizado", how="left")
    df_final = pd.merge(df_final, df_emp_folha_agrupado, on="Nome Normalizado", how="left")

    cols_zero = [
        "Férias",
        "EMPRESTIMO (EMPREGA BRASIL)",
        "EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)",
    ]
    df_final[cols_zero] = df_final[cols_zero].fillna(0.0)

    df_final["Base"] = (
        df_final["Base de Cálculo do INSS"]
        - df_final["Valor do INSS"]
        - df_final["Valor do IRRF"]
        - df_final["Férias"]
    )
    df_final["Limite de Desconto (35%)"] = df_final["Base"] * LIMITE_DESCONTO_PERCENTUAL
    df_final["Diferença (Emprega Brasil x Lista de Eventos)"] = (
        df_final["EMPRESTIMO (EMPREGA BRASIL)"]
        - df_final["EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)"]
    )
    df_final["Diferença (Limite de Desconto x Lista de Eventos)"] = (
        df_final["Limite de Desconto (35%)"]
        - df_final["EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)"]
    )

    def calcular_observacao(row):
        observacoes = []
        if abs(row["Diferença (Emprega Brasil x Lista de Eventos)"]) > 0.05:
            observacoes.append("Divergência de valores (Brasil x Folha)")
        if row["Diferença (Limite de Desconto x Lista de Eventos)"] < -0.05:
            observacoes.append("Limite 35% ultrapassado")
        return " | ".join(observacoes) if observacoes else "OK"

    df_final["Observação"] = df_final.apply(calcular_observacao, axis=1)
    df_final["Status"] = df_final["Observação"].apply(lambda obs: "Certo" if obs == "OK" else "Errado")

    df_final = df_final[COLUNAS_SAIDA].copy()
    resultados = df_final.to_dict("records")
    df_errados = df_final[df_final["Status"] == "Errado"]

    meta = {
        "total_funcionarios": len(df_final),
        "total_corretos": int((df_final["Status"] == "Certo").sum()),
        "total_errados": len(df_errados),
        "valor_divergente": float(df_errados["Diferença (Emprega Brasil x Lista de Eventos)"].abs().sum()),
        "limites_ultrapassados": int((df_final["Diferença (Limite de Desconto x Lista de Eventos)"] < -0.05).sum()),
    }
    return resultados, meta


def gerar_excel_consignados(resultados: list[dict], meta: dict | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFERÊNCIA"

    for ci, titulo in enumerate(COLUNAS_SAIDA, 1):
        cel = ws.cell(row=1, column=ci, value=titulo)
        cel.fill = COR_CABECALHO
        cel.font = FONTE_CAB
        cel.border = BORDA
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, reg in enumerate(resultados, 2):
        for ci, col in enumerate(COLUNAS_SAIDA, 1):
            valor = reg.get(col, "")
            cel = ws.cell(row=ri, column=ci, value=valor)
            cel.font = FONTE_NORMAL
            cel.border = BORDA
            cel.alignment = Alignment(vertical="center")
            if col not in ("Matrícula", "Nome do funcionário", "Status", "Observação") and isinstance(valor, (int, float)):
                cel.number_format = 'R$ #,##0.00'
                cel.alignment = Alignment(horizontal="right", vertical="center")

        status = reg.get("Status", "")
        ws.cell(row=ri, column=13).fill = COR_VERDE if status == "Certo" else COR_VERMELHO
        ws.cell(row=ri, column=13).alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 55)

    meta = meta or {}
    ws2 = wb.create_sheet("RESUMO")
    linhas = [
        ("CONFERÊNCIA DE EMPRÉSTIMOS CONSIGNADOS", "", ""),
        ("", "", ""),
        ("INDICADOR", "QUANTIDADE", "VALOR"),
        ("Total de Funcionários Processados", meta.get("total_funcionarios", len(resultados)), ""),
        ("Empregados Corretos", meta.get("total_corretos", ""), ""),
        ("Empregados com Divergência", meta.get("total_errados", ""), ""),
        ("Limites 35% Ultrapassados", meta.get("limites_ultrapassados", ""), ""),
        ("Valor Financeiro das Divergências", "", meta.get("valor_divergente", "")),
    ]

    for ri, (a, b, c_val) in enumerate(linhas, 1):
        for ci, valor in enumerate((a, b, c_val), 1):
            cel = ws2.cell(row=ri, column=ci, value=valor)
            cel.border = BORDA
            cel.font = FONTE_NORMAL
            cel.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == 3 and isinstance(valor, (int, float)):
                cel.number_format = 'R$ #,##0.00'

    ws2.merge_cells("A1:C1")
    ws2["A1"].fill = COR_CABECALHO
    ws2["A1"].font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    for ci in range(1, 4):
        ws2.cell(row=3, column=ci).fill = COR_RESUMO
        ws2.cell(row=3, column=ci).font = FONTE_BOLD

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
