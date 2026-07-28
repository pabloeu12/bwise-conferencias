"""
services/comparador.py
──────────────────────
Lógica de leitura e comparação das planilhas de Rubricas.
"""

import re
import unicodedata
import openpyxl
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Importando as funções globais e constantes do core
from core.utils import (
    moeda_para_float, limpar_str,
    COR_VERDE, COR_VERMELHO, COR_AMARELO, COR_CABECALHO, COR_RESUMO,
    FONTE_CAB, FONTE_BOLD, FONTE_NORMAL, BORDA
)

# Cor exclusiva deste módulo para o status "ausente nos lançamentos"
COR_LARANJA = PatternFill("solid", fgColor="FFD8A8")

# ── Configurações ────────────────────────────────────────────
TOLERANCIA = 0.05   # diferença máxima aceita como OK (R$)

COLUNAS_SAIDA = [
    "Matrícula", "Funcionário", "Código(s) Evento", "Nome do Evento",
    "Valor Lançamento", "Referência Sistema", "Provento Sistema",
    "Desconto Sistema", "Tipo Identificado", "Status", "Observação",
]

# ════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ════════════════════════════════════════════════════════════

def _nome_arquivo(arquivo) -> str:
    return str(getattr(arquivo, "name", getattr(arquivo, "filename", ""))).lower()

def _reposicionar(arquivo):
    try:
        arquivo.seek(0)
    except Exception:
        pass

def _carregar_matriz(arquivo) -> list[tuple]:
    """
    Lê a planilha (XLSX ou CSV) e devolve uma matriz de linhas (lista de tuplas),
    no mesmo formato que `ws.iter_rows(values_only=True)` do openpyxl produzia
    originalmente. Mantém o restante da lógica de leitura idêntico para ambos
    os formatos aceitos no uploader (.xlsx e .csv).
    """
    _reposicionar(arquivo)
    nome = _nome_arquivo(arquivo)

    if nome.endswith(".csv"):
        df = pd.read_csv(arquivo, header=None, sep=None, engine="python", dtype=object)
        df = df.where(pd.notnull(df), None)
        return [tuple(linha) for linha in df.itertuples(index=False, name=None)]

    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active
    dados = list(ws.iter_rows(values_only=True))
    wb.close()
    return dados

def extrair_codigos(nome_coluna: str) -> list[str]:
    """Extrai todos os códigos numéricos de um nome de coluna."""
    return re.findall(r"\((\d+)\)", limpar_str(nome_coluna))

def normalizar_cabecalho(valor) -> str:
    texto = limpar_str(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()

def normalizar_codigo_evento(valor) -> str:
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    texto = limpar_str(valor)
    match = re.fullmatch(r"(\d+)(?:[,.]0+)?", texto)
    return match.group(1) if match else texto

def comparar(val_lanc, ref, prov, desc) -> tuple[str, str]:
    v = moeda_para_float(val_lanc)
    if v == 0.0:
        return None, None
    r, p, d = moeda_para_float(ref), moeda_para_float(prov), moeda_para_float(desc)
    
    if abs(v - r) <= TOLERANCIA and r != 0.0:
        return "OK_REFERENCIA", "Referência"
    if abs(v - p) <= TOLERANCIA and p != 0.0:
        return "OK_PROVENTO", "Provento"
    if abs(v - d) <= TOLERANCIA and d != 0.0:
        return "OK_DESCONTO", "Desconto"
    return "DIVERGENTE", "—"

# ════════════════════════════════════════════════════════════
# LEITURA DAS PLANILHAS
# ════════════════════════════════════════════════════════════

def ler_lancamentos(arquivo) -> tuple[dict, dict, list]:
    dados = _carregar_matriz(arquivo)

    if not dados:
        raise ValueError("Planilha de lançamentos está vazia.")

    cabecalho = dados[0]
    colunas = []
    for idx, nome in enumerate(cabecalho):
        if idx < 2 or nome is None:
            continue
        codigos = extrair_codigos(nome)
        if not codigos:
            continue
        colunas.append({
            "idx":    idx,
            "nome":   limpar_str(nome),
            "codigos": codigos,
            "chave":  f"{limpar_str(nome)}||col{idx}",
            "multi":  len(codigos) > 1,
        })

    lanc_por_mat, nomes = {}, {}
    for linha in dados[1:]:
        if not linha or linha[0] is None:
            continue
        try:
            mat = int(float(str(linha[0]).strip()))
        except (ValueError, TypeError):
            continue
        nomes[mat] = limpar_str(linha[1]) if len(linha) > 1 else ""
        lanc_por_mat[mat] = {
            col["chave"]: moeda_para_float(linha[col["idx"]] if col["idx"] < len(linha) else None)
            for col in colunas
        }

    return lanc_por_mat, nomes, colunas

def ler_sistema(arquivo) -> tuple[dict, dict]:
    dados = _carregar_matriz(arquivo)

    if not dados:
        raise ValueError("Planilha do sistema está vazia.")

    inicio_dados, idxs = _mapear_colunas_sistema(dados)
    idx_func = idxs.get("func")
    sistema, nomes_sistema = {}, {}
    for linha in dados[inicio_dados:]:
        if not linha or idxs["mat"] >= len(linha) or linha[idxs["mat"]] is None:
            continue
        try:
            mat  = int(float(str(linha[idxs["mat"]]).strip()))
            cod  = normalizar_codigo_evento(linha[idxs["cod"]] if idxs["cod"] < len(linha) else None)
            nome = limpar_str(linha[idxs["nome"]]) if idxs["nome"] < len(linha) else ""
            ref  = moeda_para_float(linha[idxs["ref"]] if idxs["ref"] < len(linha) else None)
            prov = moeda_para_float(linha[idxs["prov"]] if idxs["prov"] < len(linha) else None)
            desc = moeda_para_float(linha[idxs["desc"]] if idxs["desc"] < len(linha) else None)
        except (ValueError, IndexError, TypeError):
            continue
        if not cod:
            continue

        if idx_func is not None and idx_func < len(linha) and not nomes_sistema.get(mat):
            nome_func = limpar_str(linha[idx_func])
            if nome_func:
                nomes_sistema[mat] = nome_func

        sistema.setdefault(mat, {})
        if cod in sistema[mat]:
            sistema[mat][cod]["ref"]  += ref
            sistema[mat][cod]["prov"] += prov
            sistema[mat][cod]["desc"] += desc
        else:
            sistema[mat][cod] = {"nome": nome, "ref": ref, "prov": prov, "desc": desc}

    return sistema, nomes_sistema

def _mapear_colunas_sistema(dados) -> tuple[int, dict]:
    aliases = {
        "mat":  ("matricula",),
        "cod":  ("cod evento", "codigo evento", "cod evento do recibo", "codigo evento do recibo"),
        "nome": ("evento", "nome evento", "nome do evento"),
        "ref":  ("referencia",),
        "prov": ("valor provento", "provento", "provento sistema"),
        "desc": ("valor desconto", "desconto", "desconto sistema"),
    }
    # Nome do funcionário é opcional: quando existir, serve como fonte alternativa
    # para preencher a coluna "Funcionário" de matrículas ausentes na Planilha de Lançamentos.
    aliases_func = (
        "funcionario", "nome funcionario", "nome do funcionario",
        "colaborador", "nome colaborador", "nome do colaborador",
        "empregado", "nome empregado", "nome do empregado",
        "trabalhador", "nome trabalhador", "nome do trabalhador",
        "segurado", "nome segurado", "nome do segurado",
    )

    for pos, cabecalho in enumerate(dados):
        normalizados = {
            normalizar_cabecalho(nome): idx
            for idx, nome in enumerate(cabecalho)
            if limpar_str(nome)
        }
        idxs = {}
        for campo, opcoes in aliases.items():
            idx = next((normalizados[opcao] for opcao in opcoes if opcao in normalizados), None)
            if idx is None:
                break
            idxs[campo] = idx
        else:
            idx_func = next((normalizados[opcao] for opcao in aliases_func if opcao in normalizados), None)
            if idx_func is None:
                # Cabeçalho não identifica a coluna de nome do funcionário por texto:
                # no Extrato KMM ela costuma estar na coluna C (índice 2).
                candidato = 2
                if candidato not in idxs.values():
                    idx_func = candidato
            idxs["func"] = idx_func
            return pos + 1, idxs

    if dados and len(dados[0]) >= 7:
        return 1, {"mat": 0, "func": 1, "cod": 2, "nome": 3, "ref": 4, "prov": 5, "desc": 6}

    raise ValueError("Não foi possível identificar as colunas da planilha do sistema.")

# ════════════════════════════════════════════════════════════
# COMPARAÇÃO PRINCIPAL
# ════════════════════════════════════════════════════════════

def executar_comparacao(arquivo_lanc, arquivo_sist) -> list[dict]:
    lanc_por_mat, nomes, colunas = ler_lancamentos(arquivo_lanc)
    sistema, nomes_sistema = ler_sistema(arquivo_sist)
    resultados = []

    # Códigos que aparecem em algum cabeçalho da Planilha de Lançamentos
    # (independente de o funcionário ter valor lançado ou não).
    codigos_no_cabecalho = {cod for col in colunas for cod in col["codigos"]}

    # Controla quais combinações (matrícula, código) já foram cobertas por um
    # lançamento com valor diferente de zero, para não gerar alerta duplicado
    # nem falso-positivo na verificação inversa.
    processados = set()

    for mat, valores in lanc_por_mat.items():
        nome_func   = nomes.get(mat) or nomes_sistema.get(mat, "")
        ev_sistema  = sistema.get(mat)

        for col in colunas:
            val_lanc = valores.get(col["chave"], 0.0)
            if val_lanc == 0.0:
                continue

            codigos    = col["codigos"]
            nome_event = col["nome"]

            for cod_coberto in codigos:
                processados.add((mat, cod_coberto))

            if ev_sistema is None:
                resultados.append(_linha(
                    mat, nome_func, "+".join(codigos), nome_event,
                    val_lanc, "", "", "", "—",
                    "NAO_ENCONTRADO", "Matrícula não encontrada no sistema",
                ))
                continue

            if col["multi"]:
                achados    = [c for c in codigos if c in ev_sistema]
                ref_tot    = sum(ev_sistema[c]["ref"]  for c in achados)
                prov_tot   = sum(ev_sistema[c]["prov"] for c in achados)
                desc_tot   = sum(ev_sistema[c]["desc"] for c in achados)

                if not achados:
                    status, tipo, obs = "NAO_ENCONTRADO", "—", "Nenhum dos códigos encontrado"
                else:
                    status, tipo = comparar(val_lanc, ref_tot, prov_tot, desc_tot)
                    if status is None: continue
                    obs = f"Soma de {len(achados)} eventos: {' + '.join(achados)}"

                resultados.append(_linha(
                    mat, nome_func, "+".join(codigos), nome_event,
                    val_lanc, round(ref_tot, 2), round(prov_tot, 2), round(desc_tot, 2),
                    tipo, status, obs,
                ))
                continue

            cod = codigos[0]
            if cod not in ev_sistema:
                resultados.append(_linha(
                    mat, nome_func, cod, nome_event, val_lanc, "", "", "", "—",
                    "NAO_ENCONTRADO", f"Evento {cod} não encontrado no sistema",
                ))
                continue

            ev = ev_sistema[cod]
            status, tipo = comparar(val_lanc, ev["ref"], ev["prov"], ev["desc"])
            if status is None: continue

            resultados.append(_linha(
                mat, nome_func, cod, nome_event, val_lanc,
                round(ev["ref"], 2), round(ev["prov"], 2), round(ev["desc"], 2),
                tipo, status, "",
            ))

    # ── Verificação inversa: Planilha do Sistema → Planilha de Lançamentos ──
    # Todo código com valor efetivo no sistema precisa estar previsto (e com
    # valor diferente de zero) na Planilha de Lançamentos para a mesma matrícula.
    # Só gera alerta para códigos que existem em algum título (cabeçalho) da
    # Planilha de Lançamentos — eventos que a planilha nem rastreia não entram
    # em "Ausentes nos Lançamentos".
    for mat, eventos in sistema.items():
        nome_func = nomes.get(mat) or nomes_sistema.get(mat, "")

        for cod, ev in eventos.items():
            if (mat, cod) in processados:
                continue

            if cod not in codigos_no_cabecalho:
                continue  # evento não está previsto em nenhum título da Planilha de Lançamentos

            ref, prov, desc = ev["ref"], ev["prov"], ev["desc"]
            if ref == 0.0 and prov == 0.0 and desc == 0.0:
                continue  # sem valor efetivo no sistema: não gera alerta

            if mat in lanc_por_mat:
                obs = (
                    "Evento existente na Planilha do Sistema, mas sem valor "
                    "informado na Planilha de Lançamentos para esta matrícula."
                )
            else:
                obs = (
                    "Evento existente na Planilha do Sistema, mas ausente na "
                    "Planilha de Lançamentos para esta matrícula."
                )

            resultados.append(_linha(
                mat, nome_func, cod, ev["nome"], 0.0,
                round(ref, 2), round(prov, 2), round(desc, 2),
                "—", "AUSENTE_NOS_LANCAMENTOS", obs,
            ))
            processados.add((mat, cod))

    return resultados

def _linha(mat, func, cod, evento, val, ref, prov, desc, tipo, status, obs) -> dict:
    return {
        "Matrícula":          mat,
        "Funcionário":        func,
        "Código(s) Evento":   cod,
        "Nome do Evento":     evento,
        "Valor Lançamento":   val,
        "Referência Sistema": ref,
        "Provento Sistema":   prov,
        "Desconto Sistema":   desc,
        "Tipo Identificado":  tipo,
        "Status":             status,
        "Observação":         obs,
    }

# ════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL FORMATADO
# ════════════════════════════════════════════════════════════

def gerar_excel(resultados: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFERÊNCIA"

    for ci, titulo in enumerate(COLUNAS_SAIDA, 1):
        c = ws.cell(row=1, column=ci, value=titulo)
        c.fill, c.font, c.border = COR_CABECALHO, FONTE_CAB, BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ri, reg in enumerate(resultados, 2):
        for ci, col in enumerate(COLUNAS_SAIDA, 1):
            valor = reg.get(col, "")
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font, c.border, c.alignment = FONTE_NORMAL, BORDA, Alignment(vertical="center")
            if col in ("Valor Lançamento", "Referência Sistema", "Provento Sistema", "Desconto Sistema") and isinstance(valor, float):
                c.number_format = "#,##0.00"

        status = reg.get("Status", "")
        cor = (
            COR_VERDE if "OK" in status
            else COR_VERMELHO if status == "DIVERGENTE"
            else COR_AMARELO if status == "NAO_ENCONTRADO"
            else COR_LARANJA if status == "AUSENTE_NOS_LANCAMENTOS"
            else None
        )
        if cor:
            for ci in range(1, len(COLUNAS_SAIDA) + 1): ws.cell(row=ri, column=ci).fill = cor

    ws.auto_filter.ref, ws.freeze_panes = ws.dimensions, "C2"
    larguras = [12, 38, 16, 38, 18, 18, 18, 18, 18, 18, 45]
    for ci, larg in enumerate(larguras, 1): ws.column_dimensions[get_column_letter(ci)].width = larg

    ws2 = wb.create_sheet("RESUMO")
    total = len(resultados)
    ok_ref = sum(1 for r in resultados if r["Status"] == "OK_REFERENCIA")
    ok_prov = sum(1 for r in resultados if r["Status"] == "OK_PROVENTO")
    ok_desc = sum(1 for r in resultados if r["Status"] == "OK_DESCONTO")
    ok_tot = ok_ref + ok_prov + ok_desc
    diverg = sum(1 for r in resultados if r["Status"] == "DIVERGENTE")
    nao_enc = sum(1 for r in resultados if r["Status"] == "NAO_ENCONTRADO")
    ausentes = sum(1 for r in resultados if r["Status"] == "AUSENTE_NOS_LANCAMENTOS")
    perc = round(ok_tot / total * 100, 1) if total else 0

    def _pct(qtd: int) -> str:
        return f"{round(qtd / total * 100, 1) if total else 0}%"

    # Título ocupa sempre a linha 1 (posição estrutural, não um índice "mágico").
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "CONFERÊNCIA DE FOLHA DE PAGAMENTO — MAÇANEIRO"
    ws2["A1"].fill = COR_CABECALHO
    ws2["A1"].font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    # Cada linha carrega seu próprio "tipo" de formatação — a posição na aba
    # é sempre derivada da ordem desta lista, nunca de um número fixo.
    linhas = [
        ("vazio",              "",                                 "",      ""),
        ("cabecalho",          "INDICADOR",                       "QUANTIDADE", "PERCENTUAL"),
        ("normal",             "Total comparado",                 total,    "100%"),
        ("normal",             "✅ OK — Referência",              ok_ref,   _pct(ok_ref)),
        ("normal",             "✅ OK — Provento",                ok_prov,  _pct(ok_prov)),
        ("normal",             "✅ OK — Desconto",                ok_desc,  _pct(ok_desc)),
        ("ok_total",           "✅ Total OK",                     ok_tot,   f"{perc}%"),
        ("divergente",         "❌ Divergentes",                  diverg,   _pct(diverg)),
        ("nao_encontrado",     "⚠️ Não encontrados no Sistema",   nao_enc,  _pct(nao_enc)),
        ("ausente_lancamento", "🟠 Ausentes nos Lançamentos",     ausentes, _pct(ausentes)),
        ("vazio",              "",                                 "",      ""),
        ("percentual_geral",   "PERCENTUAL DE ACERTO GERAL",      "",       f"{perc}%"),
    ]

    cores_por_tipo = {
        "cabecalho":          (COR_RESUMO,    FONTE_BOLD),
        "ok_total":           (COR_VERDE,     FONTE_BOLD),
        "divergente":         (COR_VERMELHO,  None),
        "nao_encontrado":     (COR_AMARELO,   None),
        "ausente_lancamento": (COR_LARANJA,   None),
        "percentual_geral":   (COR_CABECALHO, Font(color="FFFFFF", bold=True, name="Calibri", size=11)),
    }

    for offset, (tipo, a, b, c_val) in enumerate(linhas, start=2):
        for ci, val in enumerate((a, b, c_val), 1):
            cel = ws2.cell(row=offset, column=ci, value=val)
            cel.border = BORDA
            cel.font = FONTE_NORMAL
            cel.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        cor, fonte = cores_por_tipo.get(tipo, (None, None))
        if cor:
            for ci in range(1, 4):
                ws2.cell(row=offset, column=ci).fill = cor
                if fonte:
                    ws2.cell(row=offset, column=ci).font = fonte

    ws2.column_dimensions["A"].width, ws2.column_dimensions["B"].width, ws2.column_dimensions["C"].width = 38, 16, 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
