"""
services/consignado.py
──────────────────────
Lógica de processamento de Empréstimos Consignados (Emprega Brasil x Folha).
"""

import io
import pandas as pd
import unicodedata
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Importando constantes globais do sistema BWISE (mesmo padrão dos outros módulos)
from core.utils import (
    COR_VERDE, COR_VERMELHO, COR_CABECALHO, COR_RESUMO, 
    FONTE_CAB, FONTE_BOLD, FONTE_NORMAL, BORDA,
    moeda_para_float
)

# REGRAS DE NEGÓCIO
CODIGOS_FERIAS = [189, 190, 223]
LIMITE_DESCONTO_PERCENTUAL = 0.35
CODIGOS_EMPRESTIMO = [652, 664, 665, 666, 667, 668, 669, 677, 678, 685, 694, 695]

COLUNAS_SAIDA = [
    'Matrícula', 'Nome do funcionário', 'Base de Cálculo do INSS', 'Valor do INSS', 
    'Valor do IRRF', 'Férias', 'Base', 'Limite de Desconto (35%)', 
    'EMPRESTIMO (EMPREGA BRASIL)', 'EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)',
    'Diferença (Emprega Brasil x Lista de Eventos)', 'Diferença (Limite de Desconto x Lista de Eventos)', 
    'Status', 'Observação'
]

def normalize_name(name):
    if pd.isna(name): return ""
    name = str(name).upper()
    name = ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    return ' '.join(re.sub(r'[^A-Z\s]', '', name).split())

def load_data(file, skiprows=0, header='infer'):
    filename = str(getattr(file, 'filename', ''))
    if filename.endswith('.csv'):
        return pd.read_csv(file, skiprows=skiprows, header=header, sep=None, engine='python')
    return pd.read_excel(file, skiprows=skiprows, header=header)

def executar_conferencia_consignado(file_emprega, file_recibos, file_eventos):
    # --- 1. EMPREGA BRASIL ---
    df_emprega = load_data(file_emprega, header=0)
    col_nome_trab = 'nomeTrabalhador' if 'nomeTrabalhador' in df_emprega.columns else df_emprega.columns[df_emprega.columns.str.contains('nome', case=False)][0]
    col_valor = 'valorParcela' if 'valorParcela' in df_emprega.columns else df_emprega.columns[df_emprega.columns.str.contains('valor', case=False)][0]
    
    df_emprega['Nome Normalizado'] = df_emprega[col_nome_trab].apply(normalize_name)
    df_emprega[col_valor] = df_emprega[col_valor].apply(moeda_para_float)
    emprega_agrupado = df_emprega.groupby('Nome Normalizado', as_index=False)[col_valor].sum()
    emprega_agrupado.rename(columns={col_valor: 'EMPRESTIMO (EMPREGA BRASIL)'}, inplace=True)

    # --- 2. RECIBOS DE PAGAMENTO ---
    df_recibo = load_data(file_recibos, skiprows=2, header=None)
    df_base = pd.DataFrame()
    df_base['Matrícula'] = df_recibo[1]
    df_base['Nome do funcionário'] = df_recibo[2]
    df_base['Base de Cálculo do INSS'] = df_recibo[27].apply(moeda_para_float)
    df_base['Valor do INSS'] = df_recibo[29].apply(moeda_para_float)
    df_base['Valor do IRRF'] = df_recibo[42].apply(moeda_para_float)
    
    df_base = df_base.dropna(subset=['Matrícula', 'Nome do funcionário'])
    df_base['Nome Normalizado'] = df_base['Nome do funcionário'].apply(normalize_name)

    # --- 3. EVENTOS ---
    df_eventos = load_data(file_eventos, header=1)
    if 'Nome' not in df_eventos.columns:
        df_eventos = load_data(file_eventos, header=0)
        
    df_eventos['Nome Normalizado'] = df_eventos['Nome'].apply(normalize_name)
    if 'Valor Provento' in df_eventos.columns: df_eventos['Valor Provento'] = df_eventos['Valor Provento'].apply(moeda_para_float)
    if 'Valor Desconto' in df_eventos.columns: df_eventos['Valor Desconto'] = df_eventos['Valor Desconto'].apply(moeda_para_float)
    df_eventos['Cód. Evento Num'] = pd.to_numeric(df_eventos['Cód. Evento'], errors='coerce')
    
    df_ferias = df_eventos[df_eventos['Cód. Evento Num'].isin(CODIGOS_FERIAS)]
    df_ferias_agrupado = df_ferias.groupby('Nome Normalizado', as_index=False)['Valor Provento'].sum()
    df_ferias_agrupado.rename(columns={'Valor Provento': 'Férias'}, inplace=True)
    
    df_emp_folha = df_eventos[df_eventos['Cód. Evento Num'].isin(CODIGOS_EMPRESTIMO)].copy()
    df_emp_folha_agrupado = df_emp_folha.groupby('Nome Normalizado', as_index=False)['Valor Desconto'].sum()
    df_emp_folha_agrupado.rename(columns={'Valor Desconto': 'EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)'}, inplace=True)

    # --- 4. CRUZAMENTO (MERGE) ---
    df_final = pd.merge(df_base, df_ferias_agrupado, on='Nome Normalizado', how='left')
    df_final = pd.merge(df_final, emprega_agrupado, on='Nome Normalizado', how='left')
    df_final = pd.merge(df_final, df_emp_folha_agrupado, on='Nome Normalizado', how='left')
    
    cols_to_fill = ['Férias', 'EMPRESTIMO (EMPREGA BRASIL)', 'EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)']
    df_final[cols_to_fill] = df_final[cols_to_fill].fillna(0.0)

    df_final['Base'] = df_final['Base de Cálculo do INSS'] - df_final['Valor do INSS'] - df_final['Valor do IRRF'] - df_final['Férias']
    df_final['Limite de Desconto (35%)'] = df_final['Base'] * LIMITE_DESCONTO_PERCENTUAL
    df_final['Diferença (Emprega Brasil x Lista de Eventos)'] = df_final['EMPRESTIMO (EMPREGA BRASIL)'] - df_final['EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)']
    df_final['Diferença (Limite de Desconto x Lista de Eventos)'] = df_final['Limite de Desconto (35%)'] - df_final['EMPRESTIMO (LISTA DE EVENTOS DE RECIBO DE PAGAMENTO)']

    def avaliar_linha(row):
        obs = []
        if abs(row['Diferença (Emprega Brasil x Lista de Eventos)']) > 0.05: obs.append("Divergência de valores (Brasil x Folha)")
        if row['Diferença (Limite de Desconto x Lista de Eventos)'] < -0.05: obs.append("Limite 35% Ultrapassado")
        status = "❌ Errado" if obs else "✅ Certo"
        return pd.Series([status, " | ".join(obs) if obs else "OK"])

    df_final[['Status', 'Observação']] = df_final.apply(avaliar_linha, axis=1)

    resultados = df_final[COLUNAS_SAIDA].to_dict('records')
    
    meta = {
        "total_ativos": len(resultados),
        "total_corretos": sum(1 for r in resultados if r["Status"] == "✅ Certo"),
        "total_errados": sum(1 for r in resultados if r["Status"] == "❌ Errado"),
        "valor_divergente": df_final[df_final['Status'] == "❌ Errado"]['Diferença (Emprega Brasil x Lista de Eventos)'].abs().sum()
    }

    return resultados, meta

def gerar_excel_consignado(resultados: list[dict], meta: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFERÊNCIA CONSIGNADOS"

    # Cabeçalho
    for ci, titulo in enumerate(COLUNAS_SAIDA, 1):
        c = ws.cell(row=1, column=ci, value=titulo)
        c.fill, c.font, c.border = COR_CABECALHO, FONTE_CAB, BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Linhas
    for ri, reg in enumerate(resultados, 2):
        for ci, col in enumerate(COLUNAS_SAIDA, 1):
            valor = reg.get(col, "")
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font, c.border, c.alignment = FONTE_NORMAL, BORDA, Alignment(vertical="center")
            
            # Formatação Financeira
            if isinstance(valor, float) and col not in ["Matrícula"]:
                c.number_format = "R$ #,##0.00"
            elif col in ["Matrícula", "Status"]:
                c.alignment = Alignment(horizontal="center")

        # Coloração da Status
        status = reg.get("Status", "")
        if "Certo" in status: ws.cell(row=ri, column=13).fill = COR_VERDE
        elif "Errado" in status: ws.cell(row=ri, column=13).fill = COR_VERMELHO

    ws.auto_filter.ref, ws.freeze_panes = ws.dimensions, "C2"
    for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()